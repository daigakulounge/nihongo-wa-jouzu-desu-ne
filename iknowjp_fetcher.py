import requests
import json
from openpyxl import Workbook, load_workbook
from tools import deserial, xlsx_to_csv, downloader, make_json_list, resize_aspect_fit
from random import randint
import os, sys
from random import shuffle


isfetched = True #Have json files been downloaded?
download_bool = True # To download media change it to True
do_all_cores = False # If you want a file with all cores change it to True

path = ".//files//"
final_size = 300;

headers = {'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) ' 
					  'AppleWebKit/537.11 (KHTML, like Gecko) '
					  'Chrome/23.0.1271.64 Safari/537.11',
		'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
		'Accept-Charset': 'ISO-8859-1,utf-8;q=0.7,*;q=0.3',
		'Accept-Encoding': 'none',
		'Accept-Language': 'en-US,en;q=0.8',
		'Connection': 'keep-alive'}


# All the links to api json files

links_1 = ['https://iknow.jp/api/v2/goals/566921', 'https://iknow.jp/api/v2/goals/566922', 'https://iknow.jp/api/v2/goals/566924', 'https://iknow.jp/api/v2/goals/566925', 'https://iknow.jp/api/v2/goals/566926', 'https://iknow.jp/api/v2/goals/566927', 'https://iknow.jp/api/v2/goals/566928', 'https://iknow.jp/api/v2/goals/566929', 'https://iknow.jp/api/v2/goals/566930', 'https://iknow.jp/api/v2/goals/566932']

links_2 = ['https://iknow.jp/api/v2/goals/594768', 'https://iknow.jp/api/v2/goals/594770', 'https://iknow.jp/api/v2/goals/594771', 'https://iknow.jp/api/v2/goals/594772', 'https://iknow.jp/api/v2/goals/594773', 'https://iknow.jp/api/v2/goals/594774', 'https://iknow.jp/api/v2/goals/594775', 'https://iknow.jp/api/v2/goals/594777', 'https://iknow.jp/api/v2/goals/594778', 'https://iknow.jp/api/v2/goals/594780']

links_3 = ['https://iknow.jp/api/v2/goals/615865', 'https://iknow.jp/api/v2/goals/615866', 'https://iknow.jp/api/v2/goals/615867', 'https://iknow.jp/api/v2/goals/615869', 'https://iknow.jp/api/v2/goals/615871', 'https://iknow.jp/api/v2/goals/615872', 'https://iknow.jp/api/v2/goals/615873', 'https://iknow.jp/api/v2/goals/615874', 'https://iknow.jp/api/v2/goals/615876', 'https://iknow.jp/api/v2/goals/615877']

links_4 = ['https://iknow.jp/api/v2/goals/615947', 'https://iknow.jp/api/v2/goals/615949', 'https://iknow.jp/api/v2/goals/615950', 'https://iknow.jp/api/v2/goals/615951', 'https://iknow.jp/api/v2/goals/615953', 'https://iknow.jp/api/v2/goals/615954', 'https://iknow.jp/api/v2/goals/615955', 'https://iknow.jp/api/v2/goals/615957', 'https://iknow.jp/api/v2/goals/615958', 'https://iknow.jp/api/v2/goals/615959']

links_5 = ['https://iknow.jp/api/v2/goals/616077', 'https://iknow.jp/api/v2/goals/616078', 'https://iknow.jp/api/v2/goals/616079', 'https://iknow.jp/api/v2/goals/616080', 'https://iknow.jp/api/v2/goals/616081', 'https://iknow.jp/api/v2/goals/616082', 'https://iknow.jp/api/v2/goals/616083', 'https://iknow.jp/api/v2/goals/616084', 'https://iknow.jp/api/v2/goals/616085', 'https://iknow.jp/api/v2/goals/616086']

links_6 = ['https://iknow.jp/api/v2/goals/598434', 'https://iknow.jp/api/v2/goals/598432', 'https://iknow.jp/api/v2/goals/598431', 'https://iknow.jp/api/v2/goals/598430', 'https://iknow.jp/api/v2/goals/598427', 'https://iknow.jp/api/v2/goals/598426', 'https://iknow.jp/api/v2/goals/598425', 'https://iknow.jp/api/v2/goals/598424', 'https://iknow.jp/api/v2/goals/598423', 'https://iknow.jp/api/v2/goals/598422']

all_cores = [] 

for _i in range(1,7):
	eval('all_cores.append(links_'+str(_i) + ')') # hacky way to bring all cores together


if not isfetched:
	for links in all_cores:
		print('================================================')
		print('Core # ' + str(all_cores.index(links) + 1)) # Fetch json file
		print('================================================')

		for _link in links:

			try:
				url = _link
				fxls = str(all_cores.index(links)) + '_' + str(links.index(_link)) + '_' + url.split('/')[-1] + '.xlsx' 
				# fxls - json file name
				print(url)

				try:
					res = requests.get(url, headers=headers)
					res.raise_for_status()
				except (ConnectTimeout, HTTPError, ReadTimeout, Timeout, ConnectionError):
					res = requests.get(url, headers=headers)
					res.raise_for_status()
				data = json.loads(res.text.encode('utf-8'))
				with open(fxls + '.json', 'w') as file: #save json files localy
					json.dump(data, file)
			except Exception as e:
				print(e, e.arg)


cores = make_json_list() # fetch all the json files in the directory


if do_all_cores: #if do_all_cores = True, make csv and xlsx file with all cores
	wb = Workbook()
	ws = wb.active



for _core in cores:
	wb2 = Workbook() # xlsx for every core
	ws2 = wb2.active
	for _link in _core:
		with open(_link) as json_file:
			data = json.load(json_file)

		for first in data['goal_items']:
			outside_text_kanji, outside_text_hrkt, outside_text_eng, outside_text_part_of_speech, sound, inside_1_text_kanji, inside_1_text_kanji_blank, inside_1_text_hrkt, first_eng, first_image, first_sound, inside_2_text_kanji, inside_2_text_kanji_blank, inside_2_text_hrkt, second_eng, second_image, second_sound, distractors = deserial(first)
			# deserialize data from jsons
			
			shuffle(distractors)
			wrong1 = distractors[0]
			wrong2 = distractors[1]
			wrong3 = distractors[2]
			#download sounds and images
			filename_first_image = downloader(first_image, download_bool)
			filename_first_sound = downloader(first_sound, download_bool)

			filename_second_image = downloader(second_image, download_bool)
			filename_second_sound = downloader(second_sound, download_bool)
			
			tag = str(cores.index(_core) + 1) + 'K_core' + ' ' + outside_text_part_of_speech
			print(first_eng, second_eng,  tag, '\n')
			#make a row for csv file
			row1 = [str(randint(1,100000000)),inside_1_text_kanji_blank ,'[sound:' + filename_first_sound + ']' ,'<img src="' + filename_first_image+ '">' ,outside_text_kanji ,inside_1_text_kanji ,inside_1_text_hrkt ,first_eng , wrong1, wrong2, wrong3, tag]
			if do_all_cores:
				ws.append(row1) #write row in the all-the-cores xlsx file
			ws2.append(row1) #write row in the core xlsx file
			

			if second_eng is not 'None': #if there is a second sentense do the same for it

				row2 = [str(randint(1,100000000)),inside_2_text_kanji_blank ,'[sound:' + filename_second_sound + ']' ,'<img src="' + filename_second_image+ '">' ,outside_text_kanji ,inside_2_text_kanji ,inside_2_text_hrkt ,second_eng , wrong1, wrong2, wrong3, tag]
				if do_all_cores:
					ws.append(row2)
				ws2.append(row2)
	wb2.save(str(cores.index(_core) + 1) + 'K_core.xlsx') #save core xlsx file 
	xlsx_to_csv(str(cores.index(_core) + 1) + 'K_core.xlsx') #make core csv file 

if do_all_cores:
	wb.save('all_cores.xlsx') #save all-the-cores xlsx file 
	xlsx_to_csv("all_cores.xlsx") #make all-the-cores csv file 


resize_aspect_fit(path, final_size) #resize all images in "files" directory